"""FastAPI routes — the REST API."""

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session, selectinload
from backend.core.database import get_db
from backend.models.lead import Lead, LeadStatus, Priority
from backend.models.scrape_job import ScrapeJob, JobStatus
from backend.models.source import Source
from backend.api.schemas import (
    SourceOut, SourceToggle,
    LeadOut, LeadCreate, LeadUpdate, LeadListResponse, LeadFilters,
    BulkUpdateRequest, BulkDeleteRequest,
    ScrapeRequest, ScrapeJobOut, DashboardStats,
    TagCreate, TagOut, LeadTagUpdate,
)
from backend.services import lead_service, scrape_service
from backend.scrapers.registry import list_scrapers
from typing import Optional

router = APIRouter()


# ── Sources ──────────────────────────────────────────────

@router.get("/sources", response_model=list[SourceOut])
def get_sources(db: Session = Depends(get_db)):
    """List all scraping sources."""
    return db.query(Source).order_by(Source.name).all()


@router.patch("/sources/{source_id}", response_model=SourceOut)
def toggle_source(source_id: int, data: SourceToggle, db: Session = Depends(get_db)):
    """Enable/disable a source."""
    source = db.query(Source).filter(Source.id == source_id).first()
    if not source:
        raise HTTPException(404, "Source not found")
    source.enabled = data.enabled
    db.commit()
    db.refresh(source)
    return source


# ── Leads ────────────────────────────────────────────────

@router.get("/leads", response_model=LeadListResponse)
def get_leads(
    source: Optional[str] = None,
    priority: Optional[str] = None,
    status: Optional[str] = None,
    country: Optional[str] = None,
    lead_type: Optional[str] = None,
    search: Optional[str] = None,
    sort_by: str = "created_at",
    sort_dir: str = "desc",
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    """List leads with filtering and pagination."""
    filters = LeadFilters(
        source=source,
        priority=priority,
        status=status,
        country=country,
        lead_type=lead_type,
        search=search,
        sort_by=sort_by,
        sort_dir=sort_dir,
        page=page,
        per_page=per_page,
    )
    leads, total = lead_service.get_leads(db, filters, options=[selectinload(Lead.tags)])

    # Convert to response
    lead_items = []
    for lead in leads:
        item = LeadOut.model_validate(lead)
        item.tags = [t.name for t in lead.tags]
        lead_items.append(item)

    total_pages = (total + per_page - 1) // per_page
    return LeadListResponse(
        leads=lead_items,
        total=total,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
    )


@router.get("/leads/export")
def export_leads(
    format: str = "csv",
    source_name: Optional[str] = None,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    country: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Export leads as CSV or Excel."""
    import csv
    import io
    from fastapi.responses import StreamingResponse

    filters = LeadFilters(
        source_name=source_name, status=status, priority=priority,
        country=country, search=search, page=1, per_page=999999,
    )
    leads, _ = lead_service.get_leads(db, filters)

    if format == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=[
            "id", "company", "job_title", "source_name", "country", "city",
            "website", "status", "priority", "demand_signal", "description",
            "service_needed", "budget", "skills", "notes",
            "decision_maker_name", "decision_maker_email", "created_at",
        ])
        writer.writeheader()
        for lead in leads:
            writer.writerow({
                "id": lead.id, "company": lead.company, "job_title": lead.job_title,
                "source_name": lead.source_name, "country": lead.country, "city": lead.city,
                "website": lead.website, "status": lead.status, "priority": lead.priority,
                "demand_signal": lead.demand_signal, "description": lead.description,
                "service_needed": lead.service_needed, "budget": lead.budget,
                "skills": lead.skills, "notes": lead.notes,
                "decision_maker_name": lead.decision_maker_name,
                "decision_maker_email": lead.decision_maker_email,
                "created_at": str(lead.created_at),
            })
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=leads.csv"},
        )
    else:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import io as _io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads"

        headers = [
            "ID", "Company", "Job Title", "Source", "Country", "City",
            "Website", "Status", "Priority", "Demand Signal", "Description",
            "Service Needed", "Budget", "Skills", "Notes",
            "DM Name", "DM Email", "Created At",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for lead in leads:
            ws.append([
                lead.id, lead.company, lead.job_title, lead.source_name,
                lead.country, lead.city, lead.website, lead.status, lead.priority,
                lead.demand_signal, lead.description, lead.service_needed,
                lead.budget, lead.skills, lead.notes,
                lead.decision_maker_name, lead.decision_maker_email, str(lead.created_at),
            ])

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        output = _io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=leads.xlsx"},
        )


@router.get("/leads/filters/options")
def get_filter_options(db: Session = Depends(get_db)):
    """Get distinct values for filter dropdowns."""
    return lead_service.get_filter_options(db)


@router.post("/leads", response_model=LeadOut)
def create_lead(data: LeadCreate, db: Session = Depends(get_db)):
    """Manually create a new lead."""
    lead = lead_service.create_lead(db, data)
    item = LeadOut.model_validate(lead)
    item.tags = []
    return item


@router.patch("/leads/bulk")
def bulk_update_leads(request: BulkUpdateRequest, db: Session = Depends(get_db)):
    """Bulk update status/priority for multiple leads."""
    count = lead_service.bulk_update_leads(db, request.lead_ids, request.status, request.priority)
    return {"updated": count}


@router.delete("/leads/bulk")
def bulk_delete_leads(request: BulkDeleteRequest, db: Session = Depends(get_db)):
    """Bulk delete multiple leads."""
    count = lead_service.bulk_delete_leads(db, request.lead_ids)
    return {"deleted": count}


@router.post("/leads/import")
async def import_leads(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Import leads from a CSV file."""
    if not file.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only CSV files are supported")

    content = await file.read()
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    result = lead_service.import_leads_from_csv(db, text)
    return result


@router.get("/leads/{lead_id}", response_model=LeadOut)
def get_lead(lead_id: int, db: Session = Depends(get_db)):
    """Get a single lead."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(404, "Lead not found")
    item = LeadOut.model_validate(lead)
    item.tags = [t.name for t in lead.tags]
    return item


@router.patch("/leads/{lead_id}", response_model=LeadOut)
def update_lead(lead_id: int, data: LeadUpdate, db: Session = Depends(get_db)):
    """Update a lead."""
    lead = lead_service.update_lead(db, lead_id, data)
    if not lead:
        raise HTTPException(404, "Lead not found")
    db.commit()
    item = LeadOut.model_validate(lead)
    item.tags = [t.name for t in lead.tags]
    return item


@router.delete("/leads/{lead_id}")
def delete_lead(lead_id: int, db: Session = Depends(get_db)):
    """Delete a lead."""
    if not lead_service.delete_lead(db, lead_id):
        raise HTTPException(404, "Lead not found")
    db.commit()
    return {"ok": True}


# ── Scrape Jobs ──────────────────────────────────────────

@router.post("/scrape", response_model=ScrapeJobOut)
async def start_scrape(req: ScrapeRequest, db: Session = Depends(get_db)):
    """Start a new scrape job."""
    try:
        job = scrape_service.start_job(db, req.source_slug, req.model_dump())
        return ScrapeJobOut.model_validate(job)
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.post("/scrape/{job_id}/cancel")
async def cancel_scrape(job_id: int):
    """Cancel a running scrape job."""
    if scrape_service.cancel_job(job_id):
        return {"ok": True}
    raise HTTPException(404, "Job not found or not running")


@router.get("/scrape/jobs", response_model=list[ScrapeJobOut])
def get_scrape_jobs(
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
):
    """List recent scrape jobs."""
    jobs = (
        db.query(ScrapeJob)
        .order_by(ScrapeJob.created_at.desc())
        .limit(limit)
        .all()
    )
    return [ScrapeJobOut.model_validate(j) for j in jobs]


@router.get("/scrape/jobs/{job_id}", response_model=ScrapeJobOut)
def get_scrape_job(job_id: int, db: Session = Depends(get_db)):
    """Get details of a specific scrape job."""
    job = db.query(ScrapeJob).filter(ScrapeJob.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job not found")
    return ScrapeJobOut.model_validate(job)


# ── Dashboard Stats ──────────────────────────────────────

@router.get("/stats/trends")
def get_trends(days: int = 30, db: Session = Depends(get_db)):
    """Daily lead counts for the last N days."""
    return lead_service.get_leads_trend(db, days)


@router.get("/stats", response_model=DashboardStats)
def get_stats(db: Session = Depends(get_db)):
    """Dashboard statistics."""
    stats = lead_service.get_stats(db)
    recent_jobs = (
        db.query(ScrapeJob)
        .order_by(ScrapeJob.created_at.desc())
        .limit(5)
        .all()
    )
    stats["recent_jobs"] = [ScrapeJobOut.model_validate(j) for j in recent_jobs]
    stats["status_breakdown"] = lead_service.get_status_breakdown(db)
    return DashboardStats(**stats)


@router.get("/config")
def get_config():
    """Get current application config."""
    from backend.core.config import load_config
    cfg = load_config()
    return cfg.model_dump()


@router.put("/config")
def update_config(updates: dict, db: Session = Depends(get_db)):
    """Update application config fields."""
    from backend.core.config import load_config, save_config
    cfg = load_config()
    for key, value in updates.items():
        if hasattr(cfg, key):
            setattr(cfg, key, value)
    save_config(cfg)
    return cfg.model_dump()


# ── Tags ──────────────────────────────────────────────────

@router.get("/tags", response_model=list[TagOut])
def get_tags(db: Session = Depends(get_db)):
    """List all tags."""
    tags = lead_service.get_tags(db)
    return [TagOut.model_validate(t) for t in tags]


@router.post("/tags", response_model=TagOut)
def create_tag(data: TagCreate, db: Session = Depends(get_db)):
    """Create a new tag (or return existing if name already taken)."""
    tag = lead_service.create_tag(db, data.name, data.color or "#6366f1")
    return TagOut.model_validate(tag)


@router.delete("/tags/{tag_id}")
def delete_tag(tag_id: int, db: Session = Depends(get_db)):
    """Delete a tag by ID."""
    deleted = lead_service.delete_tag(db, tag_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Tag not found")
    return {"deleted": True}


@router.put("/leads/{lead_id}/tags")
def update_lead_tags(lead_id: int, data: LeadTagUpdate, db: Session = Depends(get_db)):
    """Replace all tags on a lead with the supplied tag IDs."""
    lead = lead_service.update_lead_tags(db, lead_id, data.tag_ids)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    item = LeadOut.model_validate(lead)
    item.tags = [t.name for t in lead.tags]
    return item
