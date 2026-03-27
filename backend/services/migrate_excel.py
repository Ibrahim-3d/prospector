"""Migrate data from the legacy Excel tracker to SQLite."""

import sys
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from openpyxl import load_workbook
from backend.core.database import init_db, SessionLocal
from backend.core.config import BASE_DIR
from backend.models.lead import Lead, Priority, LeadStatus, LeadType
from backend.models.source import Source
from backend.services.lead_service import normalize_company
from backend.services.seed import seed_sources
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

EXCEL_PATH = BASE_DIR / "ibrahim_guerrilla_targeting_v2.xlsx"

# Mapping from sheet column indices to Lead fields
TRACKER_COLUMNS = {
    2: "company",
    3: "country",
    4: "city",
    5: "category",
    8: "website",
    9: "decision_maker_name",
    10: "decision_maker_title",
    11: "decision_maker_linkedin",
    12: "decision_maker_email",
    13: "source_name",  # "How Found"
    14: "demand_signal",
    15: "service_needed",
    21: "priority",
    22: "revenue_potential",
    23: "notes",
}

PRIORITY_MAP = {
    "a+": Priority.A_PLUS,
    "a": Priority.A,
    "b": Priority.B,
    "c": Priority.C,
}


def migrate():
    """Run the Excel -> SQLite migration."""
    if not EXCEL_PATH.exists():
        logger.error("Excel file not found at %s", EXCEL_PATH)
        return

    logger.info("Initializing database...")
    init_db()
    seed_sources()

    db = SessionLocal()
    try:
        wb = load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)

        # Migrate Company Tracker
        migrated = 0
        skipped = 0

        tracker_sheet = None
        for name in wb.sheetnames:
            if "Company Tracker" in name or "📋" in name:
                tracker_sheet = wb[name]
                break

        if tracker_sheet:
            logger.info("Migrating from '%s'...", tracker_sheet.title)
            for row in tracker_sheet.iter_rows(min_row=2, values_only=False):
                cells = {i + 1: cell.value for i, cell in enumerate(row)}

                company = str(cells.get(2, "") or "").strip()
                if not company or company.lower() in ("unknown", "none", ""):
                    skipped += 1
                    continue

                norm = normalize_company(company)
                existing = db.query(Lead).filter(Lead.company_normalized == norm).first()
                if existing:
                    skipped += 1
                    continue

                priority_raw = str(cells.get(21, "A") or "A").strip().lower()
                priority = PRIORITY_MAP.get(priority_raw, Priority.A)

                source_name = str(cells.get(13, "") or "").strip()

                lead = Lead(
                    company=company,
                    company_normalized=norm,
                    lead_type=LeadType.COMPANY,
                    country=str(cells.get(3, "") or "").strip(),
                    city=str(cells.get(4, "") or "").strip(),
                    category=str(cells.get(5, "") or "").strip(),
                    website=str(cells.get(8, "") or "").strip(),
                    decision_maker_name=str(cells.get(9, "") or "").strip(),
                    decision_maker_title=str(cells.get(10, "") or "").strip(),
                    decision_maker_linkedin=str(cells.get(11, "") or "").strip(),
                    decision_maker_email=str(cells.get(12, "") or "").strip(),
                    source_name=source_name,
                    demand_signal=str(cells.get(14, "") or "").strip(),
                    service_needed=str(cells.get(15, "") or "").strip(),
                    priority=priority,
                    revenue_potential=str(cells.get(22, "") or "").strip(),
                    notes=str(cells.get(23, "") or "").strip(),
                    status=LeadStatus.NEW,
                )
                db.add(lead)
                migrated += 1

                if migrated % 100 == 0:
                    db.commit()
                    logger.info("  %d leads migrated...", migrated)

        db.commit()
        logger.info(
            "Migration complete: %d leads migrated, %d skipped (dupes/empty).",
            migrated,
            skipped,
        )

    except Exception as e:
        logger.error("Migration failed: %s", e)
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    migrate()
