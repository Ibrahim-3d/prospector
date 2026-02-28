import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import json

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_FILE = os.path.join(ROOT_DIR, 'ibrahim_guerrilla_targeting_v2.xlsx')
LOGS_DIR = os.path.join(ROOT_DIR, 'logs')
OUTPUTS_DIR = os.path.join(ROOT_DIR, 'outputs', 'json')

os.makedirs(LOGS_DIR, exist_ok=True)
os.makedirs(OUTPUTS_DIR, exist_ok=True)

# --- Configuration for Sheets ---
SHEET_CONFIGS = {
    'artstation': {
        'name': '🎨 ArtStation Jobs',
        'headers': ['#', 'Company', 'Title', 'Location', 'Posted', 'Job URL', 'Website', 'Notes', 'Priority'],
        'widths': [5, 25, 35, 25, 15, 45, 45, 50, 8]
    },
    'linkedin': {
        'name': '👔 LinkedIn Jobs',
        'headers': ['#', 'Company', 'Title', 'Location', 'Posted', 'Job URL', 'Notes', 'Priority'],
        'widths': [5, 25, 35, 25, 15, 50, 50, 8]
    },
    'wamda': {
        'name': '💰 MENA Funding',
        'headers': ['#', 'Company', 'Headline', 'Date', 'Source URL', 'Notes', 'Priority'],
        'widths': [5, 30, 60, 15, 50, 40, 8]
    },
    'upwork': {
        'name': '🔧 Upwork Projects',
        'headers': ['#', 'Job Title', 'Budget', 'Type', 'Duration', 'Experience', 'Skills', 'Description', 'Posted', 'Job URL', 'Status', 'Priority', 'Notes'],
        'widths': [5, 50, 12, 12, 18, 14, 40, 60, 16, 50, 10, 8, 30]
    }
}

def save_debug_html(page, source_name):
    """Save raw HTML for debugging scraper selectors."""
    path = os.path.join(LOGS_DIR, f'debug_{source_name}_html.html')
    content = ""
    if hasattr(page, 'html_content') and page.html_content:
        content = page.html_content
    elif hasattr(page, 'text') and page.text:
        content = page.text
    elif hasattr(page, 'body') and page.body:
        try: content = page.body.decode('utf-8', errors='ignore')
        except: content = str(page.body)
            
    with open(path, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"[DEBUG] Saved HTML dump to {path}")

def save_to_json(leads, filename):
    if not leads: return
    path = os.path.join(OUTPUTS_DIR, filename)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(leads, f, indent=4)
    print(f"[OK] Saved {len(leads)} leads to {path}")

def _ensure_sheet(wb, source_key):
    config = SHEET_CONFIGS.get(source_key)
    if not config: return None
    
    if config['name'] in wb.sheetnames:
        return wb[config['name']]
    
    ws = wb.create_sheet(config['name'])
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    
    for col, header in enumerate(config['headers'], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        
    for i, width in enumerate(config['widths'], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    return ws

def save_to_excel_structured(leads, source_key):
    """Saves leads to source-specific sheets and also the master Company Tracker."""
    if not leads: return 0
    if not os.path.exists(EXCEL_FILE): return 0
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_source = _ensure_sheet(wb, source_key)
    ws_master = wb['📋 Company Tracker'] if '📋 Company Tracker' in wb.sheetnames else None
    
    added_count = 0
    
    # Robust deduplication across both target source and master tracker
    existing_companies = set()
    
    # 1. Read existing from Source Sheet
    for r in range(2, ws_source.max_row + 1):
        val = ws_source.cell(row=r, column=2).value # 2nd column is Company/Title
        if val: existing_companies.add(str(val).lower().strip())
        
    # 2. Read existing from Master Tracker (if available)
    if ws_master and source_key != 'upwork':
        for r in range(2, ws_master.max_row + 1):
            val = ws_master.cell(row=r, column=2).value # 2nd column is Company
            if val: existing_companies.add(str(val).lower().strip())

    next_row_source = ws_source.max_row + 1
    
    for lead in leads:
        company = lead.get('company', lead.get('title', 'Unknown')).strip()
        
        # Super strict safety validation for unwanted empty outputs
        if company.lower() in existing_companies or not company or company.lower() == 'unknown':
            continue
        
        
        # 1. Save to Source-Specific Sheet
        row = next_row_source
        ws_source.cell(row=row, column=1, value=row-1) # ID
        
        if source_key == 'artstation':
            ws_source.cell(row=row, column=2, value=company)
            ws_source.cell(row=row, column=3, value=lead.get('title', ''))
            ws_source.cell(row=row, column=4, value=f"{lead.get('city', '')}, {lead.get('country', '')}")
            ws_source.cell(row=row, column=5, value=lead.get('posted', ''))
            ws_source.cell(row=row, column=6, value=lead.get('job_url', ''))
            ws_source.cell(row=row, column=7, value=lead.get('website', ''))
            ws_source.cell(row=row, column=8, value=lead.get('notes', ''))
            ws_source.cell(row=row, column=9, value=lead.get('priority', 'A'))
        elif source_key == 'linkedin':
            ws_source.cell(row=row, column=2, value=company)
            ws_source.cell(row=row, column=3, value=lead.get('title', ''))
            ws_source.cell(row=row, column=4, value=f"{lead.get('city', '')}, {lead.get('country', '')}")
            ws_source.cell(row=row, column=5, value=lead.get('posted', ''))
            ws_source.cell(row=row, column=6, value=lead.get('job_url', ''))
            ws_source.cell(row=row, column=7, value=lead.get('notes', ''))
            ws_source.cell(row=row, column=8, value=lead.get('priority', 'A'))
        elif source_key == 'wamda':
            ws_source.cell(row=row, column=2, value=company)
            ws_source.cell(row=row, column=3, value=lead.get('demand_signal', ''))
            ws_source.cell(row=row, column=4, value=lead.get('date', ''))
            ws_source.cell(row=row, column=5, value=lead.get('article_url', ''))
            ws_source.cell(row=row, column=6, value=lead.get('notes', ''))
            ws_source.cell(row=row, column=7, value=lead.get('priority', 'A'))
        elif source_key == 'upwork':
            ws_source.cell(row=row, column=2, value=lead.get('title', ''))
            ws_source.cell(row=row, column=3, value=lead.get('budget', ''))
            ws_source.cell(row=row, column=4, value=lead.get('type', ''))
            ws_source.cell(row=row, column=5, value=lead.get('duration', ''))
            ws_source.cell(row=row, column=6, value=lead.get('experience', ''))
            ws_source.cell(row=row, column=7, value=lead.get('skills', ''))
            ws_source.cell(row=row, column=8, value=lead.get('description', ''))
            ws_source.cell(row=row, column=9, value=lead.get('posted', ''))
            ws_source.cell(row=row, column=10, value=lead.get('job_url', ''))
            ws_source.cell(row=row, column=11, value="Research")
            ws_source.cell(row=row, column=12, value=lead.get('priority', 'A'))
            ws_source.cell(row=row, column=13, value=lead.get('notes', ''))

        # 2. Sync to Master Company Tracker (if it's a company lead)
        if ws_master and source_key != 'upwork':
            m_row = ws_master.max_row + 1
            ws_master.cell(row=m_row, column=1, value=m_row-1)
            ws_master.cell(row=m_row, column=2, value=company)
            ws_master.cell(row=m_row, column=3, value=lead.get('country', 'Unknown'))
            ws_master.cell(row=m_row, column=4, value=lead.get('city', ''))
            ws_master.cell(row=m_row, column=5, value=lead.get('category', 'Target'))
            ws_master.cell(row=m_row, column=8, value=lead.get('website', ''))
            ws_master.cell(row=m_row, column=13, value=lead.get('source', source_key.capitalize()))
            ws_master.cell(row=m_row, column=14, value=lead.get('demand_signal', ''))
            ws_master.cell(row=m_row, column=15, value=lead.get('service_needed', '3D/Motion'))
            ws_master.cell(row=m_row, column=20, value="Research")
            ws_master.cell(row=m_row, column=21, value=lead.get('priority', 'A'))
            ws_master.cell(row=m_row, column=23, value=lead.get('notes', ''))

        next_row_source += 1
        added_count += 1
        existing_companies.add(company.lower().strip())

    wb.save(EXCEL_FILE)
    return added_count

# Keep legacy functions for compatibility if needed, but point them to the new logic
def save_upwork_to_excel(leads): return save_to_excel_structured(leads, 'upwork')
def save_to_excel(leads):
    # This is tricky because save_to_excel in main.py receives a combined list
    # Let's split them by source if possible, or just use the master sync
    if not leads: return 0
    art_leads = [l for l in leads if l.get('source') == 'ArtStation']
    lin_leads = [l for l in leads if l.get('source') == 'LinkedIn']
    wam_leads = [l for l in leads if l.get('source') == 'Wamda']
    
    total = 0
    total += save_to_excel_structured(art_leads, 'artstation')
    total += save_to_excel_structured(lin_leads, 'linkedin')
    total += save_to_excel_structured(wam_leads, 'wamda')
    return total
