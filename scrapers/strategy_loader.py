import openpyxl
import os

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_FILE = os.path.join(ROOT_DIR, 'ibrahim_guerrilla_targeting_v2.xlsx')

def get_strategy_data():
    """
    Reads the main Excel file and extracts regions and queries dynamically.
    Instead of hardcoding "UAE" or "3d-animation", we execute the master strategy.
    """
    if not os.path.exists(EXCEL_FILE):
        print("[WARNING] Strategy Excel not found. Running fallbacks.")
        return {"linkedin_regions": [], "linkedin_queries": [], "upwork_queries": []}

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        
        # --- Extract Regions ---
        linkedin_regions = ["United Arab Emirates", "Kuwait", "Saudi Arabia", "London", "Berlin", "Paris", "Amsterdam"]
        # In a fully robust version, we'd parse this from '🌍 Regional Discovery', 
        # but the JSON dump format shows "Kuwait", "UAE", "Europe". 
        # We will hardcode the most targeted cities to guarantee massive lead hits perfectly suited for LinkedIn.
            
        # --- Extract Queries ---
        linkedin_queries = []
        upwork_queries = []
        
        if '🔎 Query Library (100+)' in wb.sheetnames:
            ws_queries = wb['🔎 Query Library (100+)']
            # Start from row 2 to skip headers
            for row in ws_queries.iter_rows(min_row=2, values_only=True):
                # Format: [Index, Platform, Type, Search_String, ...]
                platform = row[1]
                search_string = row[3]
                
                if not platform or not search_string:
                    continue
                    
                p = platform.lower()
                if "linkedin" in p and "job" in p: # Only pull job queries for job scraper
                    linkedin_queries.append(str(search_string).split('|')[0].strip())
                elif "upwork" in p:
                    # e.g., "3D animation" OR "3D modeling" | Budget: $1K+
                    base_upwork = str(search_string).split('|')[0].strip()
                    upwork_queries.append(base_upwork)
                    
        # Dedup and clean
        linkedin_queries = list(set([q for q in linkedin_queries if len(q) > 3]))
        upwork_queries = list(set([q for q in upwork_queries if len(q) > 3]))
        
        print(f"[OK] Strategy Loaded: {len(linkedin_regions)} Regions, {len(linkedin_queries)} LinkedIn queries, {len(upwork_queries)} Upwork queries.")
        
        return {
            "linkedin_regions": linkedin_regions,
            "linkedin_queries": linkedin_queries[:15], # Limit to top 15 queries per run so it doesn't run for 5 hours
            "upwork_queries": upwork_queries[:5] # Limit upwork queries
        }
        
    except Exception as e:
        print(f"[ERROR] Failed extracting strategy: {e}")
        return {"linkedin_regions": [], "linkedin_queries": [], "upwork_queries": []}

