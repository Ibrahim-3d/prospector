import json
import openpyxl
import os

def load_json_leads(filepath):
    if not os.path.exists(filepath):
        return []
    with open(filepath, "r", encoding="utf-8") as f:
        return json.load(f)

def merge_leads_to_excel():
    excel_file = 'ibrahim_guerrilla_targeting_v2.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    ws = wb['📋 Company Tracker']
    
    # Get existing companies to avoid duplicates
    existing_companies = set()
    for row in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value
        if cell_val:
            existing_companies.add(cell_val.strip().lower())
            
    all_leads = []
    all_leads.extend(load_json_leads("leads_clutch.json"))
    all_leads.extend(load_json_leads("leads_upwork.json"))
    
    if not all_leads:
        print("No new leads found in the JSON files yet. Ensure the background agents have finished.")
        return
        
    next_row = ws.max_row + 1
    added_count = 0
    
    for lead in all_leads:
        company_name = lead['company'].strip()
        if company_name.lower() in existing_companies:
            continue
            
        ws.cell(row=next_row, column=1, value=next_row - 1)
        ws.cell(row=next_row, column=2, value=company_name)
        ws.cell(row=next_row, column=3, value="Global")
        ws.cell(row=next_row, column=4, value=lead['location'])
        ws.cell(row=next_row, column=5, value=lead['title'])
        ws.cell(row=next_row, column=13, value=lead['source'])
        ws.cell(row=next_row, column=14, value=lead['demand_signal'])
        ws.cell(row=next_row, column=20, value="Research")
        ws.cell(row=next_row, column=21, value="A")
        
        existing_companies.add(company_name.lower())
        next_row += 1
        added_count += 1
        
    wb.save(excel_file)
    print(f"Successfully merged {added_count} new leads into the Excel tracker.")

if __name__ == "__main__":
    merge_leads_to_excel()
