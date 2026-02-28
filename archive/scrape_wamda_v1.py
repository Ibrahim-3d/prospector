from scrapling.fetchers import DynamicFetcher
import openpyxl
from datetime import datetime
import re

def scrape_wamda_news():
    url = "https://www.wamda.com/news"
    print(f"Scraping {url} for MENA funding signals...")
    
    page = DynamicFetcher.fetch(url)
    text = page.get_all_text()
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    
    leads = []
    
    # Wamda pattern:
    # Region (e.g. "Bahrain")
    # Title (e.g. "Foras AI raises stake...")
    # Summary
    # Date (e.g. "24 February, 2026")
    # Tags
    
    # We can search for keywords like "raises", "funding", "secures", "investment", "launches"
    keywords = ["raises", "funding", "secures", "investment", "launches", "backs", "fund"]
    
    for i, line in enumerate(lines):
        if any(kw in line.lower() for kw in keywords) and i > 0:
            # Possible headline
            headline = line
            country = lines[i-1]
            
            # Find the date (usually follows the summary)
            # Let's look ahead for a date pattern
            date = "Unknown"
            for j in range(i+1, min(i+10, len(lines))):
                if re.search(r'\d{1,2} [A-Z][a-z]+, \d{4}', lines[j]):
                    date = lines[j]
                    break
            
            # Extract company name from headline (first few words usually)
            # Heuristic: Before the keyword
            company = "Unknown"
            for kw in keywords:
                if kw in headline.lower():
                    company = headline.split(kw)[0].strip()
                    break
            
            if len(company) > 1 and len(headline) > 10:
                leads.append({
                    "company": company,
                    "country": country,
                    "headline": headline,
                    "date": date,
                    "source": "Wamda News",
                    "demand_signal": f"MENA Funding/News: {headline} ({date})"
                })
                
    return leads

def save_leads_to_excel(leads):
    filename = 'ibrahim_guerrilla_targeting_v2.xlsx'
    wb = openpyxl.load_workbook(filename)
    ws = wb['📋 Company Tracker']
    
    next_row = ws.max_row + 1
    count = 0
    
    existing_companies = set()
    for row in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value
        if cell_val:
            existing_companies.add(cell_val.strip().lower())

    for lead in leads:
        company_name = lead['company'].strip()
        # Clean up common words
        if company_name.lower() in ["uae’s", "saudi", "egypt’s", "iraq’s"]:
            # Try to get the next word or clean it
            company_name = company_name.replace("’s", "").replace("'s", "")
            
        if company_name.lower() in existing_companies:
            continue
            
        # Avoid obvious false positives
        if any(x in company_name.lower() for x in ["wamda", "search", "menu", "thought"]):
            continue

        ws.cell(row=next_row, column=1, value=next_row - 1)
        ws.cell(row=next_row, column=2, value=company_name)
        ws.cell(row=next_row, column=3, value=lead['country'])
        ws.cell(row=next_row, column=5, value="MENA Startup/Fund")
        ws.cell(row=next_row, column=13, value="Wamda")
        ws.cell(row=next_row, column=14, value=lead['demand_signal'])
        ws.cell(row=next_row, column=15, value="MENA-focused 3D/Motion Production")
        ws.cell(row=next_row, column=20, value="Research")
        ws.cell(row=next_row, column=21, value="A+")
        
        existing_companies.add(company_name.lower())
        next_row += 1
        count += 1
    
    wb.save(filename)
    return count

if __name__ == "__main__":
    leads = scrape_wamda_news()
    print(f"Total Wamda signals found: {len(leads)}")
    for l in leads[:5]:
        print(f"Found: {l['company']} in {l['country']} - {l['headline']}")
    added = save_leads_to_excel(leads)
    print(f"Added {added} new MENA leads to Excel.")
