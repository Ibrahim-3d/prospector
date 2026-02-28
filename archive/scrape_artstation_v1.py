from scrapling.fetchers import StealthyFetcher
import openpyxl
from datetime import datetime
import re

def scrape_artstation_jobs():
    url = "https://www.artstation.com/jobs/all"
    print(f"Scraping {url}...")
    
    page = StealthyFetcher.fetch(url, headless=True)
    
    # Let's try to find job listings using a more precise method
    # Since we have the text, we can see the pattern
    # Job Title
    # Company Name
    # Location
    # Posted Date
    
    jobs = []
    
    # Try to find elements that look like job listings
    # Based on ArtStation's current site, job listings are often in .job-listing class
    job_elements = page.css(".job-listing")
    print(f"Found {len(job_elements)} job elements via .job-listing selector.")
    
    if len(job_elements) == 0:
        # Fallback: parse text if selectors fail
        text = page.get_all_text()
        # Look for the pattern: 
        # Title
        # Company
        # Location
        # Posted Date
        
        # This is a bit risky but let's try to find blocks of text
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        
        # Find where "results" starts
        start_index = -1
        for i, line in enumerate(lines):
            if "results" in line.lower() and re.search(r'\d+ results', line):
                print(f"Found results at line {i}: {line}")
                start_index = i + 1
                break
        
        if start_index == -1:
             # Print first 50 lines to see where we are
             print("Could not find 'results' line. First 50 lines:")
             for idx, l in enumerate(lines[:50]):
                 print(f"{idx}: {l}")
        
        if start_index != -1:
            i = start_index
            while i < len(lines) - 4:
                # Basic heuristic for job block
                title = lines[i]
                company = lines[i+1]
                location = lines[i+2]
                posted = lines[i+3]
                
                if "Posted" in posted:
                    jobs.append({
                        "title": title,
                        "company": company,
                        "location": location,
                        "posted": posted
                    })
                    i += 4
                else:
                    i += 1
    else:
        for idx, el in enumerate(job_elements):
            try:
                # Let's try to get any text from the element to see what it is
                print(f"Element {idx} text: {el.text[:100]}")
                title = el.css_first(".job-listing-title").text.strip()
                company = el.css_first(".job-listing-company").text.strip()
                location = el.css_first(".job-listing-location").text.strip()
                posted = el.css_first(".job-listing-posted").text.strip()
                jobs.append({
                    "title": title,
                    "company": company,
                    "location": location,
                    "posted": posted
                })
            except Exception as e:
                print(f"Error parsing element {idx}: {e}")
                continue

    return jobs

def save_to_excel(jobs):
    filename = 'ibrahim_guerrilla_targeting_v2.xlsx'
    wb = openpyxl.load_workbook(filename)
    ws = wb['📋 Company Tracker']
    
    next_row = ws.max_row + 1
    count = 0
    
    # Get existing companies to avoid duplicates
    existing_companies = set()
    for row in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value
        if cell_val:
            existing_companies.add(cell_val.strip().lower())

    for job in jobs:
        company_name = job['company'].strip()
        if company_name.lower() in existing_companies:
            continue
            
        # Filter for 3D/Motion related jobs
        if not any(kw in job['title'].upper() for kw in ["3D", "MOTION", "VFX", "ANIMAT", "GENERALIST", "CGI", "UNREAL", "UNITY"]):
            continue

        # Split location into Country/City if possible
        loc_parts = job['location'].split(',')
        city = loc_parts[0].strip() if len(loc_parts) > 0 else ""
        country = loc_parts[-1].strip() if len(loc_parts) > 1 else job['location']

        ws.cell(row=next_row, column=1, value=next_row - 1) # #
        ws.cell(row=next_row, column=2, value=company_name) # Company Name
        ws.cell(row=next_row, column=3, value=country) # Country
        ws.cell(row=next_row, column=4, value=city) # City
        ws.cell(row=next_row, column=5, value="3D/CGI Studio") # Category
        ws.cell(row=next_row, column=13, value="ArtStation Jobs") # How Found
        ws.cell(row=next_row, column=14, value=f"Active job post: {job['title']}") # Demand Signal
        ws.cell(row=next_row, column=15, value="3D Animation/Motion Design") # Service They Need
        ws.cell(row=next_row, column=20, value="Research") # Status
        ws.cell(row=next_row, column=21, value="A+") # Priority
        ws.cell(row=next_row, column=23, value=f"Posted: {job['posted']}") # Notes
        
        existing_companies.add(company_name.lower())
        next_row += 1
        count += 1
    
    wb.save(filename)
    return count

if __name__ == "__main__":
    jobs = scrape_artstation_jobs()
    print(f"Total jobs found on ArtStation: {len(jobs)}")
    added = save_to_excel(jobs)
    print(f"Added {added} new leads to Excel.")
