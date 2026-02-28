from scrapling.fetchers import StealthyFetcher
import openpyxl
from datetime import datetime
import re
import time

def scrape_linkedin_jobs():
    # Search for 3D Artist jobs in UAE
    # f_TPR=r2592000 is "Past Month"
    # We want to find ones that are actually OLDER or still open after a long time
    url = "https://www.linkedin.com/jobs/search/?keywords=3d%20artist&location=United%20Arab%20Emirates&f_TPR=r2592000"
    print(f"Scraping {url} for 3D jobs in UAE...")
    
    # LinkedIn is extremely protected, using StealthyFetcher with high wait and solve_cloudflare
    page = StealthyFetcher.fetch(url, headless=True, wait=15000, solve_cloudflare=True)
    
    text = page.get_all_text()
    with open("debug_linkedin_text.txt", "w", encoding="utf-8") as f:
        f.write(text)
        
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    
    # LinkedIn Job Card Structure (Public View):
    # Job Title
    # Company Name
    # Location
    # Posted Date (e.g., "1 month ago", "3 weeks ago")
    
    leads = []
    
    # Let's try to find job blocks
    # Looking for date patterns like "weeks ago", "month ago"
    for i, line in enumerate(lines):
        if any(x in line.lower() for x in ["weeks ago", "week ago", "month ago", "months ago", "30+ days ago"]):
            try:
                # Working backwards
                # Title and Company are usually 2-4 lines above the date
                posted_date = line
                location = lines[i-1]
                company = lines[i-2]
                title = lines[i-3]
                
                # Check if it's a 3D job
                if any(kw in title.upper() for kw in ["3D", "MOTION", "CGI", "VFX", "ANIMAT"]):
                    leads.append({
                        "company": company,
                        "title": title,
                        "location": location,
                        "posted": posted_date,
                        "source": "LinkedIn Jobs",
                        "demand_signal": f"Unfilled job post: {title} ({posted_date})"
                    })
            except:
                continue
                
    return leads

def save_leads_to_excel(leads):
    filename = 'ibrahim_guerrilla_targeting_v2.xlsx'
    wb = openpyxl.load_workbook(filename)
    ws = wb['📋 Company Tracker']
    
    next_row = ws.max_row + 1
    count = 0
    
    existing_companies = set()
    for row in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value
        if cell_val:
            existing_companies.add(cell_val.strip().lower())

    for lead in leads:
        company_name = lead['company'].strip()
        
        # Avoid duplicates
        if company_name.lower() in existing_companies:
            continue
            
        # Filter out obvious UI elements
        if any(x in company_name.lower() for x in ["linkedin", "sign in", "join now", "dismiss"]):
            continue

        ws.cell(row=next_row, column=1, value=next_row - 1)
        ws.cell(row=next_row, column=2, value=company_name)
        ws.cell(row=next_row, column=3, value="UAE")
        ws.cell(row=next_row, column=4, value=lead['location'])
        ws.cell(row=next_row, column=5, value="Hiring Company")
        ws.cell(row=next_row, column=13, value="LinkedIn Jobs")
        ws.cell(row=next_row, column=14, value=lead['demand_signal'])
        ws.cell(row=next_row, column=15, value="3D Artist / Motion Design")
        ws.cell(row=next_row, column=20, value="Research")
        ws.cell(row=next_row, column=21, value="A+") # High priority for old posts
        
        existing_companies.add(company_name.lower())
        next_row += 1
        count += 1
    
    wb.save(filename)
    return count

if __name__ == "__main__":
    leads = scrape_linkedin_jobs()
    print(f"Total LinkedIn signals found: {len(leads)}")
    for l in leads:
        print(f"Found: {l['title']} at {l['company']} ({l['posted']})")
    added = save_leads_to_excel(leads)
    print(f"Added {added} new leads to Excel.")
