from scrapling.fetchers import DynamicFetcher
import openpyxl
from datetime import datetime
import re

def scrape_artstation_jobs():
    url = "https://www.artstation.com/jobs/all"
    print(f"Scraping {url}...")
    
    # Use DynamicFetcher with wait for ArtStation
    page = DynamicFetcher.fetch(url, wait=10000)
    
    jobs = []
    text = page.get_all_text()
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    
    print("First 100 lines of text:")
    for idx, line in enumerate(lines[:100]):
        print(f"{idx}: {line}")
    
    # Find all "Posted" lines
    posted_indices = []
    for i, line in enumerate(lines):
        if line.startswith("Posted "):
            posted_indices.append(i)
    
    print(f"Found {len(posted_indices)} 'Posted' lines.")
    
    for idx in posted_indices:
        try:
            posted_date = lines[idx]
            
            # Working backwards from Posted
            loc_lines = []
            j = idx - 1
            while j >= 0 and (lines[j] == "|" or lines[j].startswith("|") or "|" in lines[j] or (len(loc_lines) < 3 and "," in lines[j])):
                loc_lines.insert(0, lines[j])
                j -= 1
            
            location = " ".join(loc_lines)
            company = lines[j]
            title = lines[j-1]
            
            if len(title) > 5 and len(company) > 1:
                jobs.append({
                    "title": title,
                    "company": company,
                    "location": location,
                    "posted": posted_date
                })
            else:
                print(f"Skipping job at line {idx} due to short title/company: '{title}' at '{company}'")
        except Exception as e:
            print(f"Error parsing job at line {idx}: {e}")
            continue

    return jobs

def save_to_excel(jobs):
    filename = 'ibrahim_guerrilla_targeting_v2.xlsx'
    wb = openpyxl.load_workbook(filename)
    ws = wb['📋 Company Tracker']
    
    next_row = ws.max_row + 1
    count = 0
    
    # Get existing companies to avoid duplicates
    existing_companies = set()
    for row in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value
        if cell_val:
            existing_companies.add(cell_val.strip().lower())

    for job in jobs:
        company_name = job['company'].strip()
        # Some cleanup for company name
        if company_name.lower() in ["sign up", "sign in", "about artstation", "marketplace"]:
            continue
            
        if company_name.lower() in existing_companies:
            continue
            
        # Filter for 3D/Motion related jobs
        keywords = ["3D", "MOTION", "VFX", "ANIMAT", "GENERALIST", "CGI", "UNREAL", "UNITY", "BLENDER", "MAYA", "HOUDINI", "ZBRUSH"]
        if not any(kw in job['title'].upper() for kw in keywords):
            continue

        # Split location into Country/City if possible
        # Location might be "Malaysia | China | Japan"
        country = job['location']
        city = ""
        if "|" in job['location']:
            parts = [p.strip() for p in job['location'].split("|")]
            country = parts[0]
            if len(parts) > 1:
                city = ", ".join(parts[1:])
        elif "," in job['location']:
            parts = [p.strip() for p in job['location'].split(",")]
            city = parts[0]
            country = parts[-1]

        ws.cell(row=next_row, column=1, value=next_row - 1) # #
        ws.cell(row=next_row, column=2, value=company_name) # Company Name
        ws.cell(row=next_row, column=3, value=country) # Country
        ws.cell(row=next_row, column=4, value=city) # City
        ws.cell(row=next_row, column=5, value="3D/CGI Studio") # Category
        ws.cell(row=next_row, column=13, value="ArtStation Jobs") # How Found
        ws.cell(row=next_row, column=14, value=f"Active job post: {job['title']}") # Demand Signal
        ws.cell(row=next_row, column=15, value="3D Animation/Motion Design") # Service They Need
        ws.cell(row=next_row, column=20, value="Research") # Status
        ws.cell(row=next_row, column=21, value="A+") # Priority
        ws.cell(row=next_row, column=23, value=f"Posted: {job['posted']}") # Notes
        
        existing_companies.add(company_name.lower())
        next_row += 1
        count += 1
    
    wb.save(filename)
    return count

if __name__ == "__main__":
    jobs = scrape_artstation_jobs()
    print(f"Total jobs found on ArtStation: {len(jobs)}")
    for j in jobs:
        print(f"Found: {j['title']} at {j['company']} ({j['location']})")
    
    added = save_to_excel(jobs)
    print(f"Added {added} new leads to Excel.")
