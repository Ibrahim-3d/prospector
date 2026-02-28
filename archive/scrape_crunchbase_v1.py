from scrapling.fetchers import StealthyFetcher
import openpyxl
from datetime import datetime
import json
import re

def scrape_crunchbase_funding():
    # Crunchbase often hides data behind login, but we can try recent funding lists
    # Using a common hub URL for recent funding
    url = "https://www.crunchbase.com/lists/recent-funding-rounds"
    print(f"Scraping {url} for recent funding signals...")
    
    # Crunchbase is VERY protected, using StealthyFetcher with high wait
    page = StealthyFetcher.fetch(url, headless=True, wait=15000)
    
    text = page.get_all_text()
    with open("debug_crunchbase_text.txt", "w", encoding="utf-8") as f:
        f.write(text)
        
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    
    leads = []
    
    # We're looking for patterns like:
    # Organization Name
    # Funding Type (Seed, Series A)
    # Money Raised
    # Date
    
    # In the list view, it's often a table.
    # Let's try to find "Seed" or "Series A" markers
    for i, line in enumerate(lines):
        if any(f in line for f in ["Seed", "Series A", "Series B"]):
            try:
                # Company name is usually nearby
                # Let's look at the context
                company = lines[i-1]
                funding_type = line
                amount = lines[i+1]
                date = lines[i+2]
                
                # Basic validation
                if len(company) > 1 and "$" in amount:
                    leads.append({
                        "company": company,
                        "funding_type": funding_type,
                        "amount": amount,
                        "date": date,
                        "category": "Startup",
                        "demand_signal": f"Recent Funding: {funding_type} ({amount}) on {date}"
                    })
            except:
                continue
                
    return leads

def save_leads_to_excel(leads):
    filename = 'ibrahim_guerrilla_targeting_v2.xlsx'
    wb = openpyxl.load_workbook(filename)
    ws = wb['📋 Company Tracker']
    
    next_row = ws.max_row + 1
    count = 0
    
    existing_companies = set()
    for row in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value
        if cell_val:
            existing_companies.add(cell_val.strip().lower())

    for lead in leads:
        company_name = lead['company'].strip()
        if company_name.lower() in existing_companies:
            continue
            
        # Filter for "trash" captures
        if company_name.lower() in ["organization name", "funding round", "transaction name", "announcement date"]:
            continue

        ws.cell(row=next_row, column=1, value=next_row - 1)
        ws.cell(row=next_row, column=2, value=company_name)
        ws.cell(row=next_row, column=3, value="Unknown") # Country
        ws.cell(row=next_row, column=5, value=lead['category'])
        ws.cell(row=next_row, column=13, value="Crunchbase")
        ws.cell(row=next_row, column=14, value=lead['demand_signal'])
        ws.cell(row=next_row, column=15, value="3D/Motion for Launch/Pitch")
        ws.cell(row=next_row, column=20, value="Research")
        ws.cell(row=next_row, column=21, value="A+") # High priority for funding signals
        
        existing_companies.add(company_name.lower())
        next_row += 1
        count += 1
    
    wb.save(filename)
    return count

if __name__ == "__main__":
    leads = scrape_crunchbase_funding()
    print(f"Total potential signals found: {len(leads)}")
    added = save_leads_to_excel(leads)
    print(f"Added {added} new leads to Excel.")
